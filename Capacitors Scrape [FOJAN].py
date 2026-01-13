import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from bs4 import BeautifulSoup
import time
import json
import csv
from datetime import datetime
import random
import re
import os

class LCSCSeleniumScraperCapacitors:
    def __init__(self, headless=False):
        self.headless = headless
        self.driver = None
        self.all_products = []
        self.seen_lcsc_numbers = set()
        
    def setup_driver(self):
        """Setup Chrome driver with options"""
        chrome_options = Options()
        if self.headless:
            chrome_options.add_argument('--headless')
        
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--start-maximized')
        chrome_options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.implicitly_wait(10)
        
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
    def safe_click_next_button(self):
        """Safely click the next button using multiple methods"""
        try:
            # Method 1: Try JavaScript click first (most reliable)
            next_button = self.driver.find_element(
                By.CSS_SELECTOR, 
                'button[aria-label="Next page"]'
            )
            
            # Check if button is disabled
            if 'disabled' in next_button.get_attribute('class') or not next_button.is_enabled():
                print("Next button is disabled - reached last page")
                return False
            
            # Scroll to the button to make it visible
            self.driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
            time.sleep(0.5)
            
            # Try JavaScript click (bypasses overlay issues)
            self.driver.execute_script("arguments[0].click();", next_button)
            print("✓ Clicked next button using JavaScript")
            
            # Wait for page to load
            time.sleep(2)
            
            # Wait for products to load on new page
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "tr[id*='productId']"))
                )
                print("✓ New page loaded successfully")
                return True
            except TimeoutException:
                print("⚠️ Products didn't load after clicking next")
                return False
                
        except NoSuchElementException:
            print("❌ Could not find next page button")
            return False
        except Exception as e:
            print(f"❌ Error clicking next button: {e}")
            return False
    
    def scrape_page(self, url, max_pages=None):
        """Scrape pages using Selenium to handle JavaScript"""
        if not self.driver:
            self.setup_driver()
        
        print(f"Navigating to: {url}")
        self.driver.get(url)
        
        # Wait for page to load completely
        time.sleep(3)
        
        current_page = 1
        total_pages_scraped = 0
        
        while True:
            print(f"\n{'='*60}")
            print(f"Processing page {current_page}")
            
            # Wait for products to load
            try:
                WebDriverWait(self.driver, 15).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "tr[id*='productId']"))
                )
            except TimeoutException:
                print("❌ No products found or page didn't load properly")
                break
            
            # Get page source and parse
            page_source = self.driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            
            # Extract products
            products = self.extract_products(soup)
            self.all_products.extend(products)
            
            print(f"✓ Added {len(products)} products from page {current_page}")
            print(f"📊 Total products: {len(self.all_products)}")
            
            total_pages_scraped += 1
            
            # Check if we've reached max pages
            if max_pages and current_page >= max_pages:
                print(f"\n⏹️ Reached maximum page limit ({max_pages})")
                break
            
            # Try to go to next page
            success = self.safe_click_next_button()
            
            if not success:
                print("\n⏹️ Cannot navigate to next page, stopping")
                break
            
            current_page += 1
            
            # Random delay to be respectful
            delay = random.uniform(2, 4)
            print(f"⏳ Waiting {delay:.1f} seconds before next page...")
            time.sleep(delay)
        
        print(f"\n{'='*60}")
        print(f"🎉 Scraping complete!")
        print(f"📄 Total pages scraped: {total_pages_scraped}")
        print(f"📦 Total unique products: {len(self.all_products)}")
        
        return self.all_products
    
    def extract_products(self, soup):
        """Extract products from page"""
        products = []
        product_rows = soup.find_all('tr', id=lambda x: x and 'productId' in x)
        
        print(f"Found {len(product_rows)} product rows")
        
        for row in product_rows:
            if 'Other Suppliers' in row.get_text():
                continue
            
            product_data = self.parse_product_row(row)
            if product_data and product_data.get('Manufacturer Part Number'):
                lcsc_num = product_data['Manufacturer Part Number']
                if lcsc_num not in self.seen_lcsc_numbers:
                    self.seen_lcsc_numbers.add(lcsc_num)
                    products.append(product_data)
        
        return products
    
    def parse_product_row(self, row):
        """Parse a single capacitor product row"""
        product_data = {}
        
        # Part Number (Manufacturer Part Number)
        part_elem = row.select_one('a[href*="product-detail"][title]')
        if part_elem:
            product_data['Manufacturer Part Number'] = part_elem.get_text(strip=True)
        
        # LCSC Part Number (Supplier Part Number)
        lcsc_elem = row.select_one('a.font-Bold-600.major--text[href*="product-detail"]')
        if lcsc_elem:
            lcsc_text = lcsc_elem.get_text(strip=True)
            if lcsc_text.startswith('C'):
                product_data['Supplier Part Number'] = lcsc_text
                product_data['Link'] = 'https://www.lcsc.com' + lcsc_elem['href']
        
        if 'Manufacturer Part Number' not in product_data:
            return None
        
        # Description
        desc_elem = row.select_one('div[title].ellipsis-6')
        if desc_elem:
            product_data['description'] = desc_elem['title']
        
        cells = row.find_all('td', class_='major2--text py10')
        # Package
        package = cells[6].get_text(strip=True)
        if package and package != '-':
            product_data['Package'] = package
        
        # Capacitance
        capacitance = cells[9].get_text(strip=True)
        if capacitance and capacitance != '-':
            product_data['Capacitance'] = capacitance
        
        # Tolerance
        tolerance = cells[10].get_text(strip=True)
        if tolerance and tolerance != '-':
            product_data['Tolerance'] = tolerance

        # Voltage Rating
        voltage = cells[11].get_text(strip=True)
        if voltage and voltage != '-':
            product_data['Voltage Rating'] = voltage
        
        # Temperature Coefficient
        temp_coeff = cells[12].get_text(strip=True)
        if temp_coeff and temp_coeff != '-':
            product_data['Temperature Coefficient'] = temp_coeff
        
        # Also extract specs from description for additional data
        desc = product_data.get('description', '')
        
        # Extract capacitance from description if not already found
        if 'Capacitance' not in product_data or not product_data['Capacitance']:
            cap_patterns = [
                r'(\d+(?:\.\d+)?)\s*(pF|nF|µF|uF|mF|F)\s',
                r'(\d+(?:\.\d+)?)(pF|nF|µF|uF|mF|F)\s',
                r'(\d+(?:\.\d+)?)\s*(pf|nf|uf|mf)\b',
                r'(\d+(?:\.\d+)?)(pf|nf|uf|mf)\b'
            ]
            
            for pattern in cap_patterns:
                cap_match = re.search(pattern, desc, re.IGNORECASE)
                if cap_match:
                    product_data['Capacitance'] = f"{cap_match.group(1)}{cap_match.group(2).upper()}"
                    break
        
        # Extract voltage from description
        if 'Voltage Rating' not in product_data or not product_data['Voltage Rating']:
            volt_match = re.search(r'(\d+(?:\.\d+)?)\s*V\b', desc)
            if volt_match:
                product_data['Voltage Rating'] = f"{volt_match.group(1)}V"
        
        # Extract tolerance from description
        if 'Tolerance' not in product_data or not product_data['Tolerance']:
            tol_match = re.search(r'[±±]\s*(\d+(?:\.\d+)?)\s*%', desc)
            if tol_match:
                product_data['Tolerance'] = f"±{tol_match.group(1)}%"
        
        # Extract Temperature Coefficient (commonly X7R, X5R, Y5V, COG, etc.)
        if 'Temperature Coefficient' not in product_data or not product_data['Temperature Coefficient']:
            # Common temperature coefficient codes
            temp_coeff_patterns = [
                r'(X7R|X5R|Y5V|Z5U|X8R|X6S|X7S|X7T|X7U)',
                r'(COG|NPO)',
                r'(SL|XH|X7M)'
            ]
            
            for pattern in temp_coeff_patterns:
                temp_match = re.search(pattern, desc, re.IGNORECASE)
                if temp_match:
                    product_data['Temperature Coefficient'] = temp_match.group(1).upper()
                    break
            
            # If still not found, check for temperature coefficient in description
            if 'Temperature Coefficient' not in product_data:
                # Look for temperature range patterns
                temp_range_match = re.search(r'([+-]\d+(?:\.\d+)?°C\s*(?:to|~)\s*[+-]\d+(?:\.\d+)?°C)', desc)
                if temp_range_match:
                    product_data['Temperature Coefficient'] = temp_range_match.group(1)
        
        return product_data
    
    def save_to_json(self, filename=None):
        """Save products to JSON file"""
        if not self.all_products:
            print("No products to save")
            return False
        
        if not filename:
            filename = f'Capacitors-FOJAN.json'
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.all_products, f, indent=2, ensure_ascii=False)
            print(f"✓ Saved {len(self.all_products)} products to JSON: {filename}")
            return True
        except Exception as e:
            print(f"✗ Error saving JSON: {e}")
            return False
    
    def save_to_csv(self, filename=None):
        """Save products to CSV file"""
        if not self.all_products:
            print("No products to save")
            return False
        
        if not filename:
            filename = f'Capacitors-FOJAN.csv'
        
        # Updated fields list
        fields = [
            'Manufacturer Part Number',
            'Supplier Part Number',
            'Description',
            'Package',
            'Capacitance',
            'Tolerance',
            'Voltage Rating',
            'Temperature Coefficient',
            'Link'
        ]
        
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fields)
                writer.writeheader()
                for product in self.all_products:
                    row_data = {
                        'Manufacturer Part Number': product.get('Manufacturer Part Number', ''),
                        'Supplier Part Number': product.get('Supplier Part Number', ''),
                        'Description': product.get('description', ''),
                        'Package': product.get('Package', ''),
                        'Capacitance': product.get('Capacitance', ''),
                        'Tolerance': product.get('Tolerance', ''),
                        'Voltage Rating': product.get('Voltage Rating', ''),
                        'Temperature Coefficient': product.get('Temperature Coefficient', ''),
                        'Link': product.get('Link', '')
                    }
                    writer.writerow(row_data)
            print(f"✓ Saved {len(self.all_products)} products to CSV: {filename}")
            return True
        except Exception as e:
            print(f"✗ Error saving CSV: {e}")
            return False
    
    def save_to_excel(self, filename=None):
        """Save products to Excel file"""
        if not self.all_products:
            print("No products to save to Excel")
            return False
        
        if not filename:
            filename = f'Capacitors-FOJAN.xlsx'
        
        try:
            # Try using pandas first (recommended)
            import pandas as pd
            
            # Convert to DataFrame
            df = pd.DataFrame(self.all_products)
            
            # Rename 'description' to 'Description' for better column name
            if 'description' in df.columns:
                df.rename(columns={'description': 'Description'}, inplace=True)
            
            # Updated column order
            column_order = [
                'Manufacturer Part Number',
                'Supplier Part Number',
                'Description',
                'Package',
                'Capacitance',
                'Tolerance',
                'Voltage Rating',
                'Temperature Coefficient',
                'Link'
            ]
            
            # Reorder columns (keep any additional columns at the end)
            existing_cols = [col for col in column_order if col in df.columns]
            other_cols = [col for col in df.columns if col not in column_order]
            df = df[existing_cols + other_cols]
            
            # Save to Excel
            df.to_excel(filename, index=False)
            print(f"✓ Saved {len(self.all_products)} products to Excel: {filename}")
            return True
            
        except ImportError:
            # Fall back to openpyxl if pandas is not available
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                
                wb = Workbook()
                ws = wb.active
                ws.title = "LCSC Capacitors"
                
                # Updated headers
                headers = [
                    'Manufacturer Part Number',
                    'Supplier Part Number',
                    'Description',
                    'Package',
                    'Capacitance',
                    'Tolerance',
                    'Voltage Rating',
                    'Temperature Coefficient',
                    'Link'
                ]
                
                # Write headers
                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_num, value=header)
                
                # Write data
                for row_num, product in enumerate(self.all_products, 2):
                    ws.cell(row=row_num, column=1, value=product.get('Manufacturer Part Number', ''))
                    ws.cell(row=row_num, column=2, value=product.get('Supplier Part Number', ''))
                    ws.cell(row=row_num, column=3, value=product.get('description', ''))
                    ws.cell(row=row_num, column=4, value=product.get('Package', ''))
                    ws.cell(row=row_num, column=5, value=product.get('Capacitance', ''))
                    ws.cell(row=row_num, column=6, value=product.get('Tolerance', ''))
                    ws.cell(row=row_num, column=7, value=product.get('Voltage Rating', ''))
                    ws.cell(row=row_num, column=8, value=product.get('Temperature Coefficient', ''))
                    ws.cell(row=row_num, column=9, value=product.get('Link', ''))
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save workbook
                wb.save(filename)
                print(f"✓ Saved {len(self.all_products)} products to Excel: {filename}")
                return True
                
            except ImportError:
                print("✗ Excel export requires either pandas or openpyxl.")
                print("  Install with: pip install pandas openpyxl")
                return False
        
        except Exception as e:
            print(f"✗ Error saving Excel: {e}")
            return False
    
    def save_all_formats(self, base_filename=None):
        """Save to all formats (JSON, CSV, Excel)"""
        if not self.all_products:
            print("No products to save")
            return False
        
        if not base_filename:
            base_filename = f'Capacitors-FOJAN'
        
         # Define folder structure
        base_output_folder = "Outputs"
        json_folder = os.path.join(base_output_folder, "JSONs")
        csv_folder = os.path.join(base_output_folder, "CSVs")
        excel_folder = os.path.join(base_output_folder, "Excels")

        # Create folders if they don't exist
        for folder in [base_output_folder, json_folder, csv_folder, excel_folder]:
            if not os.path.exists(folder):
                os.makedirs(folder)
                print(f"Created folder: {folder}")




        print(f"\n{'='*60}")
        print("SAVING TO ALL FORMATS")
        print('='*60)
        
        results = []
        
        # Save to JSON
        json_file = os.path.join(json_folder, f'{base_filename}.json')
        json_success = self.save_to_json(json_file)
        results.append(('JSON', json_file, json_success))
        
        # Save to CSV
        csv_file = os.path.join(csv_folder, f'{base_filename}.csv')
        csv_success = self.save_to_csv(csv_file)
        results.append(('CSV', csv_file, csv_success))
        
        # Save to Excel
        excel_file = os.path.join(excel_folder, f'{base_filename}.xlsx')
        excel_success = self.save_to_excel(excel_file)
        results.append(('Excel', excel_file, excel_success))
        
        # Print summary
        print(f"\n{'='*60}")
        print("SAVE RESULTS SUMMARY")
        print('='*60)
        
        for format_name, filename, success in results:
            status = "✓ SUCCESS" if success else "✗ FAILED"
            rel_path = os.path.relpath(filename, os.getcwd())
            print(f"{format_name:6} : {status} - {rel_path}")
        
        return all([success for _, _, success in results])
    
    def display_summary(self):
        """Display summary of scraped products"""
        if not self.all_products:
            print("No products to display")
            return
        
        print("\n" + "="*60)
        print(f"CAPACITORS SUMMARY")
        print("="*60)
        print(f"Total unique products: {len(self.all_products)}")
        
        # Field completion statistics
        fields_to_check = [
            'Manufacturer Part Number',
            'Supplier Part Number',
            'description',
            'Package',
            'Capacitance',
            'Tolerance',
            'Voltage Rating',
            'Temperature Coefficient',
            'Link'
        ]
        
        print("\nField completion statistics:")
        for field in fields_to_check:
            count = sum(1 for p in self.all_products if p.get(field))
            percentage = (count / len(self.all_products)) * 100
            print(f"  {field:25} {count:4}/{len(self.all_products):4} ({percentage:5.1f}%)")
        
        if self.all_products:
            print("\nFirst 3 capacitors:")
            print("-"*60)
            for i, product in enumerate(self.all_products[:3], 1):
                print(f"\n{i}. {product.get('Manufacturer Part Number', 'N/A')}")
                print(f"   LCSC: {product.get('Supplier Part Number', 'N/A')}")
                print(f"   Description: {product.get('description', 'N/A')[:80]}...")
                print(f"   Package: {product.get('Package', 'N/A')}")
                print(f"   Specs: {product.get('Capacitance', 'N/A')} | "
                      f"{product.get('Voltage Rating', 'N/A')} | "
                      f"{product.get('Tolerance', 'N/A')}")
                print(f"   Temp Coefficient: {product.get('Temperature Coefficient', 'N/A')}")
    
    def close(self):
        """Close the driver"""
        if self.driver:
            self.driver.quit()

def main():
    """Main function"""
    print("="*60)
    print("LCSC CAPACITOR Scraper - FOJAN Brand")
    print("="*60)
    
    # Check for required Excel libraries
    print("\nChecking for Excel export capabilities...")
    excel_available = False
    try:
        import pandas
        print("✓ pandas library available for Excel export")
        excel_available = True
    except ImportError:
        try:
            from openpyxl import Workbook
            print("✓ openpyxl library available for Excel export")
            excel_available = True
        except ImportError:
            print("✗ Neither pandas nor openpyxl found.")
            print("  Excel export will not be available.")
            print("  Install with: pip install pandas openpyxl")
            sys.exit()
    print("\n" + "="*60)
    

    try:
        # Ask if user wants to run in headless mode
        headless_input = input("Run in background (headless mode)? (y/n, default=n): ").strip().lower()
        headless = headless_input == 'y'
        
        scraper = LCSCSeleniumScraperCapacitors(headless=headless)
        
        # Capacitor URL (FOJAN brand)
        url = 'https://www.lcsc.com/category/1142.html?brand=13046'

        print("Try to scrape all pages")
        print("\nAttempting to scrape all pages (this may take a while)...")
        
        # Start scraping
        products = scraper.scrape_page(url, None) # Replace None with a number to limit pages
        
        if products:
            print(f"\n✅ Successfully scraped {len(products)} capacitors")
            
            # Display summary
            scraper.display_summary()
            
            # Save options
            base_filename = f'Capacitors-FOJAN'
            scraper.save_all_formats(base_filename)        
            print("\n✅ Export completed!")
            print("="*60)
        else:
            print("\n❌ No capacitors were scraped")
        
        # Ask if user wants to keep browser open
        keep_open = input("\nKeep browser window open for inspection? (y/n, default=n): ").strip().lower()
        if keep_open != 'y':
            scraper.close()
            print("Browser closed.")
        else:
            print("Browser window will remain open. Close it manually when done.")
        
    except ImportError:
        print("\n❌ Selenium not installed.")
        print("Install with: pip install selenium")
        print("Also download ChromeDriver from: https://chromedriver.chromium.org/")
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        try:
            scraper.close()
        except:
            pass

if __name__ == "__main__":
    main()
    print("\n" + "="*60)
    print("Program completed successfully!")
    print("="*60)